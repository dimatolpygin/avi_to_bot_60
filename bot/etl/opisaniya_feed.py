# -*- coding: utf-8 -*-
"""Индекс ОПИСАНИЙ объявлений Авито из фида автозагрузки (этап 14, «подтянуть описание»).

Мессенджер Авито в `context.value` даёт только заголовок и цену объявления, а ТЕКСТ
описания (где лежат размеры, материал, состав, объём) через API недоступен вовсе
(проверено вживую: Item Core API отдаёт только status/url/时间, описания нет). Единственный
структурный источник описаний — ФИД АВТОЗАГРУЗКИ, который аккаунт сам выгружает из кабинета
Авито (xlsx: лист на категорию, строки-объявления). Ключ объявления в фиде — колонка
«Номер объявления на Авито» (= `AvitoId`), и он СОВПАДАЕТ с `context.value.id` из мессенджера
(проверено на соль/трубе/окне) — поэтому по item_id из чата описание находится напрямую.

Этот модуль разбирает фид в компактный индекс `data/opisaniya_avito.json`
(`{item_id: {"title", "opisanie"}}`), который ядро грузит при старте и подмешивает в факт
объявления. Описание САНИРУЕТСЯ: срезаются HTML, телефоны, «☎ звоните» и разделители-заборы —
и чтобы бот не попугайничал номер в чат (контакты только когда клиент сам спросил), и чтобы
публичный репозиторий не тянул телефоны. Остаётся продуктовая суть (размеры/материал/польза).

Рефреш ручной: заказчик выгружает свежий фид → `python -m bot.etl.opisaniya_feed <файл.xlsx>`
→ пуш → автодеплой (прод монтирует `./data:/app/data:ro`). Скрейпинг avito.ru НЕ делаем.
"""
from __future__ import annotations

import argparse
import html
import json
import os
import re

try:
    from ..logger import logger
except Exception:  # запуск как одиночный скрипт вне пакета
    import logging
    logger = logging.getLogger("opisaniya_feed")

# data/ в корне проекта: bot/etl/opisaniya_feed.py → bot/etl → bot → корень.
KATALOG_DANNYH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data")
FAYL_PO_UMOLCHANIYU = os.path.join(KATALOG_DANNYH, "opisaniya_avito.json")

# Колонки фида приходят и на русском (товарные листы), и на английском (Работа-Вакансии).
IMENA_AVITOID = {"Номер объявления на Авито", "AvitoId"}
IMENA_TITLE = {"Название объявления", "Title"}
IMENA_DESCR = {"Описание объявления", "Description"}

# Потолок длины санированного описания. Медиана исходников ~1.4к символов, но продуктовая
# суть (размеры/материал) идёт в начале; 1000 хватает и держит промпт в узде.
LIMIT_OPISANIYA = 1000

_TELEFON = re.compile(r"(?:\+?7|8)[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d"
                      r"[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d[\s\-()]*\d")
_ZABOR = re.compile(r"^[\s=~_\-*·•.]{4,}$")
_CTA = re.compile(r"звоните|делайте заказ|пишите в личку|заказывайте прямо", re.I)
_TEL_ZNAKI = ("☎", "📞", "📲", "✆", "whatsapp", "ватсап", "вотсап", "viber", "telegram")


def _strip_html(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"</p\s*>", "\n", s, flags=re.I)
    s = re.sub(r"<li[^>]*>", "\n• ", s, flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)
    return html.unescape(s)


def _sanirovat(raw) -> str:
    """HTML → чистый текст; выкинуть телефоны, «☎ звоните», заборы-разделители; подрезать."""
    text = _strip_html(raw)
    stroki: list[str] = []
    for ln in text.splitlines():
        ln = re.sub(r"[ \t ]+", " ", ln).strip()
        if not ln or _ZABOR.match(ln):
            continue
        low = ln.lower()
        if _TELEFON.search(ln) or _CTA.search(ln) or any(z in low for z in _TEL_ZNAKI):
            continue
        stroki.append(ln)
    out = "\n".join(stroki).strip()
    out = re.sub(r"\n{2,}", "\n", out)
    if len(out) > LIMIT_OPISANIYA:
        obrez = out[:LIMIT_OPISANIYA].rsplit(" ", 1)[0].rstrip(" ,;.-")
        out = obrez + "…"
    return out


def _kolonka(hdr: dict, imena: set[str]):
    for n in imena:
        if n in hdr:
            return hdr[n]
    return None


def _kak_id(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v))
    return str(v).strip()


def razobrat_feed(put: str) -> dict[str, dict]:
    """Фид автозагрузки (.xlsx) → `{item_id: {"title", "opisanie"}}`.

    Идём по всем листам-категориям (кроме «Инструкция» и справочников «Спр-…»),
    шапка — строка 2, данные — с 5-й. Ключ — «Номер объявления на Авито». Дубль ключа
    (одно объявление в двух листах — так не бывает, но подстрахуемся) → WARNING, берём
    последний, молча не проглатываем.
    """
    import openpyxl

    # read_only на этих фидах не отдаёт строки (сбитые dimensions листа) — грузим обычным
    # режимом; 1.2 МБ / ~1.4к строк для памяти не проблема.
    wb = openpyxl.load_workbook(put, data_only=True)
    index: dict[str, dict] = {}
    dublej = 0
    bez_opisaniya = 0
    for name in wb.sheetnames:
        if name == "Инструкция" or name.startswith("Спр-"):
            continue
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)              # строка 1 — хлебные крошки категории
            hdr_row = next(rows)    # строка 2 — имена колонок
            next(rows)              # строка 3 — обязательность
            next(rows)              # строка 4 — подсказка
        except StopIteration:
            continue
        hdr = {str(v).strip(): i for i, v in enumerate(hdr_row) if v is not None}
        c_av = _kolonka(hdr, IMENA_AVITOID)
        c_ti = _kolonka(hdr, IMENA_TITLE)
        c_de = _kolonka(hdr, IMENA_DESCR)
        if c_av is None or c_de is None:
            logger.warning("Фид: лист «%s» без колонок AvitoId/Описание — пропущен", name)
            continue
        for r in rows:
            if c_av >= len(r):
                continue
            item_id = _kak_id(r[c_av])
            if not item_id:
                continue
            title = str(r[c_ti]).strip() if c_ti is not None and c_ti < len(r) and r[c_ti] else ""
            opisanie = _sanirovat(r[c_de]) if c_de < len(r) else ""
            if not opisanie:
                bez_opisaniya += 1
            if item_id in index:
                dublej += 1
            index[item_id] = {"title": title, "opisanie": opisanie}
    wb.close()
    logger.info("Фид разобран: объявлений %d, дублей ключа %d, без описания %d",
                len(index), dublej, bez_opisaniya)
    return index


def sohranit_json(index: dict[str, dict], out: str) -> None:
    with open(out, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=0, sort_keys=True)
    logger.info("Индекс описаний записан: %s (%d объявлений)", out, len(index))


def zagruzit_indeks(katalog: str = KATALOG_DANNYH) -> dict[str, str]:
    """Слить все `data/opisaniya*.json` в плоский `{item_id: opisanie}` для рантайма.

    item_id уникален по всему Авито, поэтому несколько файлов (по аккаунтам) сливаются
    без коллизий. Нет файлов — пустой словарь (фича просто не «загорается»), не падаем.
    """
    import glob

    plosky: dict[str, str] = {}
    for put in sorted(glob.glob(os.path.join(katalog, "opisaniya*.json"))):
        try:
            with open(put, encoding="utf-8") as f:
                dannye = json.load(f)
        except (OSError, ValueError) as e:
            logger.warning("Индекс описаний «%s» не прочитан: %s", put, e)
            continue
        for item_id, zap in dannye.items():
            opis = (zap.get("opisanie") or "").strip() if isinstance(zap, dict) else ""
            if opis:
                plosky[str(item_id)] = opis
    if plosky:
        logger.info("Описаний объявлений загружено: %d", len(plosky))
    return plosky


def main() -> None:
    ap = argparse.ArgumentParser(description="Разбор фида автозагрузки Авито в индекс описаний")
    ap.add_argument("fayl", help="Путь к .xlsx фида автозагрузки")
    ap.add_argument("--out", default=FAYL_PO_UMOLCHANIYU, help="Куда записать JSON-индекс")
    ap.add_argument("--tolko-razbor", action="store_true", help="Разобрать и показать статистику, не писать файл")
    args = ap.parse_args()

    index = razobrat_feed(args.fayl)
    print(f"Объявлений с описанием: {sum(1 for v in index.values() if v['opisanie'])} / {len(index)}")
    if not args.tolko_razbor:
        sohranit_json(index, args.out)
        print(f"Записано: {args.out}")


if __name__ == "__main__":
    main()
