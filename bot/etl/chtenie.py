# -*- coding: utf-8 -*-
"""Чтение прайса клиента и валидация ДО похода в БД.

Источник правды — выгрузка гугл-таблицы заказчика **«Инфа по стройке для бота»**,
лист Saunamart: 141 строка, колонки `article, nomenclature, characteristics,
price_apiece, price_per_meter, Наличие`. Её заказчик ведёт руками и обновляет.

⚠️ **Не путать со старым `Копия ИП Семенко ЕВ.xlsx`** (тоже лежит в `материалы/`).
Это рабочая свалка на 13 листов, куда смотрел старый бот: там 229 строк, другие
артикулы (`23405` против `123405`), другие цены (дверь 5609 против 11700), нет
колонки наличия и вдобавок четыре разных среза одного ассортимента. Смешение
этих срезов — и есть причина бага «бот выдаёт разные цены».

Формат xlsx тоже читается (параметр `list_name` — имя листа): гугл-таблицу
могут выгрузить и так.

Главное правило модуля: **до записи в БД файл должен быть проверен целиком**.
Битый или обрезанный прайс не должен ополовинить каталог — лучше упасть
с русским объяснением и оставить в БД вчерашние данные.
"""
from __future__ import annotations

import csv
import hashlib
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

from ..logger import logger
from .razbor import normalizovat

# Заголовки, без которых читать нечего.
OBYAZATELNYE_KOLONKI = ("article", "nomenclature", "price_apiece")
# Колонка наличия. Заказчик ведёт её сам: «Да» = есть, пусто = уточнить
# у менеджера. Значения «Нет» в таблице не бывает.
KOLONKA_NALICHIYA = ("наличие", "в наличии", "availability", "stock")

# Заголовок → каноничный ключ колонки. Два диалекта одного прайса: CSV-выгрузка
# несёт англоимена (`price_apiece`), живая Google-таблица — русские (`цена штука`).
# Поддерживаем оба, чтобы фолбэк на CSV и основной Google-источник читались одной
# читалкой. Ключ `price_per_m2` («цена за м2») — новая колонка курса 31.07:
# распознаём здесь, а в `StrokaPraysa` её проносит этап A2.
ALIASY_KOLONOK = {
    "article": "article", "артикул": "article",
    "nomenclature": "nomenclature", "номенклатура": "nomenclature",
    "наименование": "nomenclature",
    "characteristics": "characteristics", "характеристики": "characteristics",
    "price_apiece": "price_apiece", "цена штука": "price_apiece",
    "цена за штуку": "price_apiece", "цена шт": "price_apiece",
    "price_per_meter": "price_per_meter", "цена м.п.": "price_per_meter",
    "цена мп": "price_per_meter", "цена за м.п.": "price_per_meter",
    "price_per_m2": "price_per_m2", "цена за м2": "price_per_m2",
    "цена за м²": "price_per_m2", "цена м2": "price_per_m2",
}

# Заголовок в выгрузке гугл-таблицы стоит не первой строкой: сверху пустая
# строка, а слева пустая колонка. Ищем строку с `article` в первых десяти.
STROK_ISKAT_ZAGOLOVOK = 10

# Меньше этого числа строк — файл почти наверняка обрезан. В прайсе 141 строка;
# порог 50 отсекает аварию, но переживёт сокращение ассортимента вдвое.
MINIMUM_STROK = 50


class OshibkaPraysa(Exception):
    """Файл прайса непригоден. Текст сообщения идёт человеку как есть."""


@dataclass(frozen=True)
class StrokaPraysa:
    """Одна строка файла, приведённая к типам. Ещё без разбора названия."""

    article: str
    name: str
    characteristics: str | None
    price_apiece: Decimal | None
    price_per_meter_fayl: Decimal | None   # только для сверки инварианта, в БД не идёт
    nalichie_syroe: str | None             # None = колонки в файле нет
    nomer_stroki: int
    # Цена за м² прямо из колонки таблицы (курс 31.07). None = колонки нет или
    # ячейка пуста → цену за м² вычислим фолбэком. В файловом CSV колонки обычно
    # нет, у живой Google-таблицы — есть.
    price_per_m2_fayl: Decimal | None = None


@dataclass(frozen=True)
class Prays:
    """Прочитанный и проверенный файл целиком."""

    stroki: list[StrokaPraysa]
    istochnik: str
    data_praysa: datetime
    vsego_strok: int


# ── Приведение ячеек ─────────────────────────────────────────────────────────

def _artikul(v) -> str | None:
    """Артикул из ячейки. openpyxl отдаёт числовые артикулы как float
    (`23405.0`) — иначе ключ синхронизации разъедется между запусками."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = normalizovat(v)
    # То же самое, когда файл пришёл CSV: строка «23405.0».
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


def _cena(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        s = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".")
        c = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return c if c > 0 else None


def _klyuch_bez_artikula(nazvanie: str) -> str:
    """Синтетический ключ для позиции без артикула — «мягкий ассортимент»
    заказчика (электрокаменки ЭКМ без SKU, курс 31.07). Артикул нужен как ключ
    синхронизации (uq account+article), а у этих строк его нет. Берём стабильный
    хеш номенклатуры: пока название в таблице не меняется, ключ тот же — повторный
    синк такие позиции не задваивает. Префикс `nomen-` виден в БД и логах, сразу
    ясно, что артикул не из файла, а выведен по названию."""
    h = hashlib.sha1(nazvanie.encode("utf-8")).hexdigest()[:12]
    return f"nomen-{h}"


# ── Чтение ───────────────────────────────────────────────────────────────────

def _nayti_zagolovok(syrye: list[list], put: str) -> int:
    """Номер строки с заголовками. В выгрузке гугл-таблицы сверху пустая строка,
    поэтому «заголовок — первая строка» тут не работает."""
    for i, stroka in enumerate(syrye[:STROK_ISKAT_ZAGOLOVOK]):
        if any(v is not None and normalizovat(v).lower() == "article" for v in stroka):
            return i
    raise OshibkaPraysa(
        f"❌ В прайсе «{put}» не нашёл строку заголовков: ни в одной из первых "
        f"{STROK_ISKAT_ZAGOLOVOK} строк нет колонки `article`.\n"
        f"   Проверь, что выгружен нужный лист таблицы «Инфа по стройке для бота»."
    )


def _indeksy_kolonok(zagolovok: list, put: str) -> dict[str, int]:
    """Сопоставить заголовки файла с нужными колонками. Нет обязательной —
    падаем сразу, до чтения строк."""
    naydeno: dict[str, int] = {}
    for i, h in enumerate(zagolovok):
        if h is None:
            continue
        klyuch = normalizovat(h).lower()
        if klyuch in ALIASY_KOLONOK:
            naydeno.setdefault(ALIASY_KOLONOK[klyuch], i)
        elif klyuch in KOLONKA_NALICHIYA:
            naydeno.setdefault("nalichie", i)

    net = [k for k in OBYAZATELNYE_KOLONKI if k not in naydeno]
    if net:
        raise OshibkaPraysa(
            f"❌ В прайсе «{put}» нет обязательных колонок: {', '.join(net)}.\n"
            f"   Найдены: {', '.join(sorted(naydeno)) or '— ни одной'}.\n"
            f"   Нужен лист с заголовками article / nomenclature / price_apiece."
        )
    return naydeno


def _syrye_stroki_xlsx(put: str, list_name: str | None) -> list[list]:
    try:
        import openpyxl
    except ImportError as e:                                   # pragma: no cover
        raise OshibkaPraysa(f"❌ Не установлен openpyxl, читать xlsx нечем: {e}")

    try:
        wb = openpyxl.load_workbook(put, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — сюда попадает и «файл не xlsx», и битый архив
        raise OshibkaPraysa(f"❌ Не удалось открыть прайс «{put}»: {e}")

    try:
        if list_name is None:
            list_name = wb.sheetnames[0]
        elif list_name not in wb.sheetnames:
            raise OshibkaPraysa(
                f"❌ В прайсе «{put}» нет листа «{list_name}».\n"
                f"   Листы в файле: {', '.join(wb.sheetnames)}."
            )
        return [list(r) for r in wb[list_name].iter_rows(values_only=True)]
    finally:
        wb.close()


def _syrye_stroki_csv(put: str) -> list[list]:
    # utf-8-sig: выгрузка из Excel почти всегда с BOM.
    with open(put, encoding="utf-8-sig", newline="") as f:
        obrazec = f.read(4096)
        f.seek(0)
        try:
            dialekt = csv.Sniffer().sniff(obrazec, delimiters=",;\t")
        except csv.Error:
            dialekt = csv.excel
        return [list(r) for r in csv.reader(f, dialekt)]


def sobrat_prays(syrye: list[list], istochnik: str, data_praysa: datetime,
                 opisanie: str) -> Prays:
    """Собрать и проверить `Prays` из сырых строк листа (CSV, xlsx или Google).

    Единое ядро для всех источников: находит заголовок, сопоставляет колонки,
    приводит типы, отбрасывает хвост и битые строки, проверяет минимум. `opisanie`
    подставляется в тексты ошибок (имя файла или «Google-таблица, лист …»), а
    `istochnik`/`data_praysa` источник заполняет сам — у файла это mtime, у
    таблицы — момент чтения.
    """
    if not syrye:
        raise OshibkaPraysa(f"❌ Прайс «{opisanie}» пуст — ни одной строки.")

    nomer_zagolovka = _nayti_zagolovok(syrye, opisanie)
    kolonki = _indeksy_kolonok(syrye[nomer_zagolovka], opisanie)
    est_nalichie = "nalichie" in kolonki
    logger.info("📄 Прайс «%s»: заголовок в строке %d, колонка наличия %s",
                istochnik, nomer_zagolovka + 1,
                "есть" if est_nalichie else "ОТСУТСТВУЕТ — всё пойдёт как «наличие неизвестно»")

    def yacheyka(stroka: list, klyuch: str):
        i = kolonki.get(klyuch)
        return stroka[i] if i is not None and i < len(stroka) else None

    stroki: list[StrokaPraysa] = []
    bez_ceny = bez_nazvaniya = bez_artikula = 0
    for nomer, syraya in enumerate(syrye[nomer_zagolovka + 1:], start=nomer_zagolovka + 2):
        art = _artikul(yacheyka(syraya, "article"))
        nazvanie = yacheyka(syraya, "nomenclature")
        nazvanie = normalizovat(nazvanie) if nazvanie else ""
        if not art and not nazvanie:
            continue                                   # хвост пустых строк листа
        if not nazvanie:
            bez_nazvaniya += 1
            logger.warning("📄 Строка %d: артикул %s без названия — пропускаю", nomer, art)
            continue
        cena = _cena(yacheyka(syraya, "price_apiece"))
        # Строка без артикула — «мягкий ассортимент» (ЭКМ без SKU): ключим по
        # номенклатуре и держим даже без цены («есть в наличии, цену уточню»).
        # У строки С артикулом пустая цена — это порча данных, её пропускаем.
        if not art:
            art = _klyuch_bez_artikula(nazvanie)
            bez_artikula += 1
        elif cena is None:
            bez_ceny += 1
            logger.warning("📄 Строка %d: «%s» без цены — пропускаю", nomer, nazvanie[:60])
            continue
        har = yacheyka(syraya, "characteristics")
        nal = yacheyka(syraya, "nalichie") if est_nalichie else None
        stroki.append(StrokaPraysa(
            article=art,
            name=nazvanie,
            characteristics=normalizovat(har) if har else None,
            price_apiece=cena,
            price_per_meter_fayl=_cena(yacheyka(syraya, "price_per_meter")),
            nalichie_syroe=("" if nal is None else normalizovat(nal)) if est_nalichie else None,
            nomer_stroki=nomer,
            price_per_m2_fayl=_cena(yacheyka(syraya, "price_per_m2")),
        ))

    if bez_nazvaniya or bez_ceny:
        logger.warning("📄 Пропущено строк: без названия %d, без цены %d",
                       bez_nazvaniya, bez_ceny)
    if bez_artikula:
        logger.info("📄 Позиций без артикула (ключ по номенклатуре): %d", bez_artikula)

    if len(stroki) < MINIMUM_STROK:
        raise OshibkaPraysa(
            f"❌ В прайсе «{opisanie}» пригодных строк {len(stroki)}, "
            f"а должно быть не меньше {MINIMUM_STROK}.\n"
            f"   Похоже, источник обрезан или это не тот лист. В БД ничего не меняю."
        )

    logger.info("📄 Прочитано строк: %d, дата прайса: %s",
                len(stroki), data_praysa.strftime("%d.%m.%Y %H:%M"))
    return Prays(stroki=stroki, istochnik=istochnik,
                 data_praysa=data_praysa, vsego_strok=len(stroki))


def prochitat(put: str, list_name: str | None = None) -> Prays:
    """Прочитать файл и проверить его целиком. Любая беда → `OshibkaPraysa`."""
    if not os.path.isfile(put):
        raise OshibkaPraysa(
            f"❌ Файл прайса не найден: «{put}».\n"
            f"   Свежую выгрузку клади в материалы/прайс/ — она монтируется в контейнер."
        )

    rasshirenie = os.path.splitext(put)[1].lower()
    syrye = (_syrye_stroki_csv(put) if rasshirenie == ".csv"
             else _syrye_stroki_xlsx(put, list_name))
    data = datetime.fromtimestamp(os.path.getmtime(put), tz=timezone.utc)
    return sobrat_prays(syrye, os.path.basename(put), data, put)


# ── Дубли артикулов ──────────────────────────────────────────────────────────

def shlopnut_dubli(stroki: list[StrokaPraysa]) -> tuple[list[StrokaPraysa], int, int]:
    """Свести строки к одной на артикул. Возвращает (строки, дублей, конфликтов).

    В файле клиента 25 дублей, и все **полные**: артикул, название и цена
    совпадают, дублируется вся строка (печи ЭКМ, фольга, камни, пульты).
    Такой дубль схлопывается без потери данных — ровно поэтому ключом
    синхронизации выбран артикул.

    Если однажды всплывёт дубль с РАЗНЫМИ ценами — это уже поломка прайса,
    молча её проглатывать нельзя: пишем WARNING с обеими ценами и берём
    последнюю строку (она ниже в файле, обычно свежее).
    """
    po_artikulu: dict[str, StrokaPraysa] = {}
    dubley = konfliktov = 0
    for s in stroki:
        bylo = po_artikulu.get(s.article)
        if bylo is not None:
            dubley += 1
            if bylo.price_apiece != s.price_apiece or bylo.name != s.name:
                konfliktov += 1
                logger.warning(
                    "📄 Артикул %s повторяется с разными данными: строка %d «%s» %s ₽ "
                    "против строки %d «%s» %s ₽ — беру последнюю",
                    s.article, bylo.nomer_stroki, bylo.name[:40], bylo.price_apiece,
                    s.nomer_stroki, s.name[:40], s.price_apiece)
        po_artikulu[s.article] = s

    if dubley:
        logger.info("📄 Дублей артикула схлопнуто: %d (из них с расхождением: %d), "
                    "позиций осталось: %d", dubley, konfliktov, len(po_artikulu))
    return list(po_artikulu.values()), dubley, konfliktov
